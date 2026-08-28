import asyncio
import io
import re
import uuid
from dataclasses import dataclass
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine
from sqlalchemy.pool import NullPool

from backend.modules.wb_core.application import SellerNotFoundError
from backend.modules.wb_core.infrastructure.postgres import SellerRepository
from backend.modules.wb_reviews.domain import DailyRatings
from backend.modules.wb_reviews.infrastructure.postgres import ReviewSyncRepository

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass(frozen=True, slots=True)
class ReviewReportFile:
    seller_name: str
    date_from: date
    date_to: date
    content: bytes

    @property
    def filename(self) -> str:
        slug = re.sub(r"[^\w-]+", "-", self.seller_name, flags=re.UNICODE).strip("-") or "seller"
        return f"reviews_{slug}_{self.date_from.isoformat()}_{self.date_to.isoformat()}.xlsx"


class ReviewReportService:
    """Собирает xlsx-отчёт по снапшотам отзывов селлера за период."""

    def __init__(self, database_url: str) -> None:
        self._database_url = database_url

    async def build(self, seller_id: uuid.UUID, date_from: date, date_to: date) -> ReviewReportFile:
        return await asyncio.to_thread(self._build, seller_id, date_from, date_to)

    def _build(self, seller_id: uuid.UUID, date_from: date, date_to: date) -> ReviewReportFile:
        # Весь отчёт — от чтения из БД до сборки книги — живёт в этом треде.
        # Соединения общего движка привязаны к event loop приложения, поэтому
        # тред поднимает собственный одноразовый движок без пула.
        seller_name, names, counts = asyncio.run(self._load(seller_id, date_from, date_to))
        content = self._render(seller_name, names, counts, date_from, date_to)
        return ReviewReportFile(seller_name, date_from, date_to, content)

    async def _load(
        self, seller_id: uuid.UUID, date_from: date, date_to: date
    ) -> tuple[str, dict[str, str], list[DailyRatings]]:
        engine = create_async_engine(self._database_url, poolclass=NullPool)
        try:
            async with async_sessionmaker(engine, expire_on_commit=False)() as session:
                sellers = SellerRepository(session)
                seller = await sellers.get(seller_id)
                if seller is None:
                    raise SellerNotFoundError
                names = {article.article: article.name for article in await sellers.list_articles(seller_id)}
                counts = await ReviewSyncRepository(session).history_range(seller_id, date_from, date_to)
                return seller.name, names, counts
        finally:
            await engine.dispose()

    @staticmethod
    def _render(
        seller_name: str,
        names: dict[str, str],
        counts: list[DailyRatings],
        date_from: date,
        date_to: date,
    ) -> bytes:
        days = [date_from + timedelta(offset) for offset in range((date_to - date_from).days + 1)]
        by_article: dict[str, dict[date, int]] = {}
        for count in counts:
            by_article.setdefault(count.article, {})[count.date] = sum(count.ratings)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Отзывы"

        header_fill = PatternFill("solid", fgColor="1F2A44")
        header_font = Font(bold=True, color="FFFFFF")
        centered = Alignment(horizontal="center", vertical="center")

        sheet.append(["Артикул", "Название"] + [day.strftime("%d.%m.%Y") for day in days])
        for column in range(1, len(days) + 3):
            cell = sheet.cell(row=1, column=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = centered

        for article in sorted(by_article):
            snapshots = by_article[article]
            # Пустая ячейка означает «в этот день снапшота нет», а не ноль отзывов.
            row = [article, names.get(article, "—")] + [snapshots.get(day) for day in days]
            sheet.append(row)
            for column in range(3, len(days) + 3):
                sheet.cell(row=sheet.max_row, column=column).alignment = centered

        sheet.column_dimensions["A"].width = 14
        sheet.column_dimensions["B"].width = 42
        for index in range(len(days)):
            sheet.column_dimensions[get_column_letter(3 + index)].width = 12
        sheet.freeze_panes = "C2"

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
