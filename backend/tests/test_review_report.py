import io
from datetime import date

from openpyxl import load_workbook
from sqlalchemy import delete

from backend.modules.wb_core.infrastructure.postgres.models import ArticleModel, SellerModel
from backend.modules.wb_reviews.application import ReviewReportFile, ReviewReportService
from backend.modules.wb_reviews.infrastructure.postgres import ReviewSyncRepository
from backend.modules.wb_reviews.infrastructure.postgres.models import DailyReviewCountModel
from backend.shared.settings import load_settings
from backend.storage.pg import Database


def test_report_filename_keeps_period_and_sanitizes_name() -> None:
    report = ReviewReportFile("ООО «Ромашка» / основной", date(2026, 8, 1), date(2026, 8, 28), b"")
    assert report.filename == "reviews_ООО-Ромашка-основной_2026-08-01_2026-08-28.xlsx"


async def test_report_builds_snapshot_matrix_from_postgres() -> None:
    settings = load_settings("backend/shared/settings/config.test.yaml")
    database = Database()
    await database.connect(settings.database.url, pool_size=1, max_overflow=0)
    seller = SellerModel(name="Report seller", catalog_sync_status="success")
    date_from, date_to = date(2026, 8, 1), date(2026, 8, 3)

    try:
        async with database.session() as session:
            session.add(seller)
            await session.flush()
            session.add(ArticleModel(seller_id=seller.id, article="111", name="Коврик XL"))
            reviews = ReviewSyncRepository(session)
            # 2 августа снапшота нет — в отчёте должна остаться пустая ячейка.
            await reviews.upsert_daily_counts(seller.id, date(2026, 8, 1), {"111": (1, 0, 0, 0, 4)})
            await reviews.upsert_daily_counts(seller.id, date(2026, 8, 3), {"111": (1, 0, 0, 1, 5)})
            # Соседние даты за пределами периода в отчёт попадать не должны.
            await reviews.upsert_daily_counts(seller.id, date(2026, 8, 4), {"111": (9, 9, 9, 9, 9)})
            await session.commit()

        report = await ReviewReportService(settings.database.url).build(seller.id, date_from, date_to)

        sheet = load_workbook(io.BytesIO(report.content)).active
        assert [cell.value for cell in sheet[1]] == [
            "Артикул",
            "Название",
            "01.08.2026",
            "02.08.2026",
            "03.08.2026",
        ]
        assert [cell.value for cell in sheet[2]] == ["111", "Коврик XL", 5, None, 7]
        assert sheet.max_row == 2
        assert report.filename == "reviews_Report-seller_2026-08-01_2026-08-03.xlsx"
    finally:
        async with database.session() as session:
            await session.execute(delete(DailyReviewCountModel).where(DailyReviewCountModel.seller_id == seller.id))
            await session.execute(delete(ArticleModel).where(ArticleModel.seller_id == seller.id))
            await session.execute(delete(SellerModel).where(SellerModel.id == seller.id))
            await session.commit()
        await database.disconnect()
